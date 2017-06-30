#! python3

import requests
import json
import re
import time
import openpyxl
import itchat

def getVV(name, id):
	url1 = 'http://v.youku.com/action/getVideoPlayInfo?beta&timestamp=&vid=';	#js请求前部
	url2 = '&showid=0&param[]=share&param[]=favo&param[]=download&param[]=phonewatch&param[]=updown&callback=tuijsonp4';	#js请求后部
	url = url1 + id + url2;
	#print(id[i]);
	res = requests.get(url);
	jsonRegex = re.compile(r'{.*}');	#从请求响应中正则出json报文
	mo = jsonRegex.search(res.text);
	#print(mo.group());
	jsonData = json.loads(mo.group());	#str转为dict
	data = jsonData.get('data', -1);
	stat = data.get('stat', -1)
	vv = stat.get('vv', -1);
	print(name + ':' + vv);
	
	return vv;

def sleeptime(hour, min, sec):
    return hour*3600 + min*60 + sec;

if __name__ == "__main__":

	#itchat.auto_login();
	#user = itchat.search_friends(name=u'C')[0];
	#user.send(u'开始监控流量');
	#time.sleep(2)
	#user.send('初始流量为:');
      
	#视频ID
	id = {'高博文':'696095143',
			'包一峰':'692981281',
			'姜鹏':'691444866',
			'叶泳湘':'685103558',
			'马薇薇':'678648839',
			'郭培':'669683458',
			'缪歌':'662895723'
			};

	wb = openpyxl.Workbook();	#创建工作簿
	#新建工作表并初始化表头
	for k,v in id.items():
		sheet = wb.create_sheet(k);
		sheet.cell(row=1, column=1).value = '时间';
		sheet.cell(row=1, column=2).value = '播放量';

	currentTime = time.strftime('%Y-%m-%d%H%M',time.localtime(time.time()));

	#保存文件，根据当前时间
	wbName = 'C:/Users/Chan/Desktop/log/log_' + currentTime + '.xlsx';
	wb.save(wbName); 
	
	rows = 2;	#起始行
	pauseTime = sleeptime(0,10,0);	#程序暂停时间

	for i in range(500):
		print(time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(time.time())));
		for k,v in id.items(): 
			#print(k,v);
			VV = getVV(k, v)
			wb[k].cell(row=rows, column=1).value = time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(time.time()));
			wb[k].cell(row=rows, column=2).value = VV;
			time.sleep(1)
			''''
			numRegex = re.compile('[0-9]*')
			mo = numRegex.findall(VV)
			num = ''
			for j in range(len(mo)):
				num += mo[j]
			#print(num)
			if int(num) > 600000 or int(num)<10000:
				msg = k + ':' + VV
				user.send('报警!\n' + msg);
				time.sleep(2);
			if (0 == i):
				msg = k + ':' + VV
				user.send(msg);
				time.sleep(2);
			'''
		rows += 1;
		wb.save(wbName);
		time.sleep(pauseTime);



