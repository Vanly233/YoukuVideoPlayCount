#! python3
#*--coding=utf8--*

import requests
import json
import re
import time
import openpyxl

def getVV(name, id):
	url = 'http://cache.video.qiyi.com/jp/pc/' + id +'/'
	headers_chrome = {
		'User-Agent': 'Mozilla/5.0 (Windows NT 6.3; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/57.0.2987.133 Safari/537.36'}
	res = requests.get(url, headers_chrome)
	jsonRegex = re.compile(r'{.*}')
	mo = jsonRegex.search(res.text)
	# print(mo.group());
	jsonData = json.loads(mo.group())
	vv = jsonData.get(id, -1)
	print(name + ':' + str(vv));
	return str(vv)

def sleeptime(hour, min, sec):
    return hour*3600 + min*60 + sec;

def main():
	id = {
		'严歌苓': '722633000',
		'周平': '717466900',
		'高博文': '685756100',
		'包一峰': '681226900',
		'姜鹏': '679046100',
		'叶泳湘': '671149200',
		'马薇薇': '656368000',
		'郭培': '643499700'
	}

	wb = openpyxl.Workbook()

	for k,v in id.items():
		sheet = wb.create_sheet(k)
		sheet.cell(row=1, column=1).value = '时间'
		sheet.cell(row=1, column=2).value = '播放量'

	currentTime = time.strftime('%Y-%m-%d%H%M', time.localtime(time.time()));

	# 保存文件，根据当前时间
	wbName = 'C:/Users/Chan/Desktop/temp/qiyi/log_' + currentTime + '.xlsx';
	wb.save(wbName);

	rows = 2;  # 起始行
	pauseTime = sleeptime(0, 30, 0);  # 程序暂停时间

	for i in range(500):
		print(time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(time.time())));
		for k,v in id.items():
			#print(k,v);
			VV = getVV(k, v)
			wb[k].cell(row=rows, column=1).value = time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(time.time()));
			wb[k].cell(row=rows, column=2).value = VV;
			time.sleep(1)
		rows += 1;
		wb.save(wbName);
		time.sleep(pauseTime);

if __name__ == '__main__':
	main()
