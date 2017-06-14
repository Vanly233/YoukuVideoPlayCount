#! python3

import requests
import json
import re
import time
import openpyxl

'''
id = ['XMjc4NDM4MDU3Mg',		#高博文
		'XMjc3MTkyNTEyNA',		#包一峰
		'XMjc2NTc3OTQ2NA',		#姜鹏
		'XMjc0MDQxNDIzMg',		#叶泳湘
		'XMjcxNDU5NTM1Ng'		#马薇薇	
];
'''

def getVV(name, id):
	url1 = 'http://v.youku.com/action/getVideoPlayInfo?beta&timestamp=&vid=';		#js请求前部
	url2 = '&showid=0&param[]=share&param[]=favo&param[]=download&param[]=phonewatch&param[]=updown&callback=tuijsonp4';	#js请求后部
	url = url1 + id + url2;
	#print(id[i]);
	res = requests.get(url);
	jsonRegex = re.compile(r'{.*}');	#从请求响应中正则出json报文
	mo = jsonRegex.search(res.text);
	#print(mo.group());
	jsonData = json.loads(mo.group());	#str转为dict
	data = jsonData.get('data', -1);
	stat = data.get('stat', -1);
	vv = stat.get('vv', -1);
	print(name + ':' + vv);
	
	return vv;

def sleeptime(hour,min,sec):
    return hour*3600 + min*60 + sec;

if __name__ == "__main__":
	#视频ID
	id = {'高博文':'696095143',
			'包一峰':'692981281',
			'姜鹏':'691444866',
			'叶泳湘':'685103558',
			'马薇薇':'678648839'
			};

	wb = openpyxl.Workbook();	#创建工作簿
	#新建工作表并初始化表头
	for k,v in id.items():
		sheet = wb.create_sheet(k);
		sheet.cell(row=1, column=1).value = '时间';
		sheet.cell(row=1, column=2).value = '播放量';
	wb.save('E:/log.xlsx'); 

	rows = 2;
	pause = sleeptime(0,10,0);
	for i in range(100):
		print(time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(time.time())));
		for k,v in id.items(): 
			#print(k,v);
			wb[k].cell(row=rows, column=1).value = time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(time.time()));
			wb[k].cell(row=rows, column=2).value = getVV(k, v);
		rows += 1;
		wb.save('E:/log.xlsx'); 	
		time.sleep(pause);



